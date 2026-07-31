# -*- coding: utf-8 -*-
"""Генерирует src/data/clients.ts из Excel-файла с данными клиентов."""
from __future__ import annotations

import json
import os
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "подсчет (1).xlsx"
OUT = Path(__file__).resolve().parents[1] / "src" / "data" / "clients.ts"

STATUS_MAP = {
    "лояльный": "Лояльный",
    "нейтральный": "Нейтральный",
    "критик": "Критик",
    "нет данных": "Нет данных",
}

LOYALTY_BY_ZONE = {
    "critical": [
        "Приоритетное участие в конференции Альфа-Банка",
        "Начислить баллы по программе лояльности",
        "Предложить недостающие продукты из АППП",
        "Оформить льготу на ПУ",
    ],
    "neutral": [
        "Приоритетное участие в конференции Альфа-Банка",
        "Начислить баллы по программе лояльности",
        "Предложить недостающие продукты из АППП",
        "Индивидуальные условия на ПУ",
    ],
    "loyal": ["Предложить недостающие продукты из АППП"],
    "unknown": ["Предложить продукты из кошелька клиента (АППП)"],
}


def ts_str(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def zone_from_status(status: str, index: float | None) -> str:
    if status == "Нет данных" or index is None:
        return "unknown"
    if status == "Критик" or index <= 3:
        return "critical"
    if status == "Лояльный" or index >= 5:
        return "loyal"
    return "neutral"


def parse_claim_details(claim: str) -> dict | None:
    lower = claim.lower()
    if not claim.strip():
        return None
    if "не выявлен" in lower or "претензий нет" in lower or "претензий со стороны клиента нет" in lower:
        return None
    if "претенз" not in lower and "обращен" not in lower:
        return None

    status = "Закрыта"
    if "открыт" in lower and "закрыт" not in lower:
        status = "Открыта"

    theme = "Претензия клиента"
    theme_match = re.search(r"Темы претензий[^–\-:]+[–\-]\s*([^.;]+)", claim, re.I)
    if theme_match:
        theme = theme_match.group(1).strip()
    elif "комисси" in lower and "тариф" in lower:
        theme = "Комиссии и тарифы"

    resolution = claim.split("Инструмент повышения лояльности:")[0].strip()
    if "закрыт" in lower:
        closed_match = re.search(
            r"((?:обе )?закрыт[^.;]*|возврат[^.;]*|отклонен[^.;]*)",
            claim,
            re.I,
        )
        if closed_match:
            resolution = closed_match.group(1).strip()

    reason = "См. описание в ClaimCRM"
    reason_match = re.search(r"причин[аы][^:]*[:\-–]\s*([^.;]+)", claim, re.I)
    if reason_match:
        reason = reason_match.group(1).strip()

    details: dict = {
        "theme": theme,
        "status": status,
        "reason": reason,
        "resolution": resolution,
    }

    comp_match = re.search(r"возврат[^.;]*", claim, re.I)
    if comp_match:
        details["compensation"] = comp_match.group(0).strip().rstrip(".")

    return details


def extract_loyalty_tips(claim: str, zone: str) -> list[str]:
    match = re.search(r"Инструмент повышения лояльности:\s*(.+)", claim, re.I | re.S)
    if match:
        tip = match.group(1).strip().rstrip(".")
        if tip:
            return [tip]
    return LOYALTY_BY_ZONE[zone]


def load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        pin = ws.cell(r, 1).value
        if not pin:
            continue
        raw_index = ws.cell(r, 2).value
        pulse_index = float(raw_index) if raw_index is not None else None
        raw_status = str(ws.cell(r, 3).value or "").strip().lower()
        status = STATUS_MAP.get(raw_status, "Нет данных")
        rows.append(
            {
                "pin": str(pin).strip(),
                "pulseIndex": pulse_index,
                "status": status,
                "briefAssessment": str(ws.cell(r, 5).value or "").strip(),
                "positiveSides": str(ws.cell(r, 6).value or "").strip(),
                "negativeSides": str(ws.cell(r, 7).value or "").strip(),
                "claimSummary": str(ws.cell(r, 8).value or "").strip(),
                "recommendation": str(ws.cell(r, 9).value or "").strip(),
            }
        )
    return rows


def render_client(row: dict) -> str:
    zone = zone_from_status(row["status"], row["pulseIndex"])
    claim_details = parse_claim_details(row["claimSummary"])
    loyalty_tips = extract_loyalty_tips(row["claimSummary"], zone)

    index_literal = "null" if row["pulseIndex"] is None else str(row["pulseIndex"])

    extras: list[str] = []
    if claim_details:
        lines = ["      claimDetails: {"]
        for key, val in claim_details.items():
            lines.append(f"        {key}: {ts_str(val)},")
        lines.append("      },")
        extras.extend(lines)
    if loyalty_tips != LOYALTY_BY_ZONE[zone]:
        tips = ",\n".join(f"        {ts_str(t)}" for t in loyalty_tips)
        extras.append(f"      loyaltyTips: [\n{tips},\n      ],")

    extras_block = ""
    if extras:
        extras_block = ",\n    {\n" + "\n".join(extras) + "\n    }"

    return f"""  makeClient(
    {ts_str(row['pin'])},
    {index_literal},
    {ts_str(row['status'])},
    {ts_str(row['briefAssessment'])},
    {ts_str(row['positiveSides'])},
    {ts_str(row['negativeSides'])},
    {ts_str(row['claimSummary'])},
    {ts_str(row['recommendation'])}{extras_block},
  )"""


def main() -> None:
    rows = load_rows()
    if not rows:
        raise SystemExit("No rows found in xlsx")

    clients_body = ",\n".join(render_client(r) for r in rows)

    content = f"""import type {{ ClientPulse }} from '../types';
import {{ getZoneFromStatus }} from '../types';

const loyaltyByZone = {{
  critical: [
    'Приоритетное участие в конференции Альфа-Банка',
    'Начислить баллы по программе лояльности',
    'Предложить недостающие продукты из АППП',
    'Оформить льготу на ПУ',
  ],
  neutral: [
    'Приоритетное участие в конференции Альфа-Банка',
    'Начислить баллы по программе лояльности',
    'Предложить недостающие продукты из АППП',
    'Индивидуальные условия на ПУ',
  ],
  loyal: ['Предложить недостающие продукты из АППП'],
  unknown: ['Предложить продукты из кошелька клиента (АППП)'],
}};

function makeClient(
  pin: string,
  pulseIndex: number | null,
  status: ClientPulse['status'],
  briefAssessment: string,
  positiveSides: string,
  negativeSides: string,
  claimSummary: string,
  recommendation: string,
  extras?: Partial<ClientPulse>,
): ClientPulse {{
  const zone = getZoneFromStatus(status, pulseIndex);
  return {{
    pin,
    companyName: pin,
    inn: 'XXXXXXXXXX',
    pulseIndex,
    status,
    zone,
    briefAssessment,
    positiveSides,
    negativeSides,
    claimSummary,
    recommendation,
    tenureYears: 3,
    tenureMonths: 1,
    tenureDays: 0,
    vocScores: [],
    suggestedProducts: [],
    loyaltyTips: loyaltyByZone[zone],
    ...extras,
  }};
}}

export const clients: ClientPulse[] = [
{clients_body}
];

export function getClientByPin(pin: string): ClientPulse | undefined {{
  return clients.find((c) => c.pin === pin);
}}

export function searchClients(query: string): ClientPulse[] {{
  const q = query.trim().toLowerCase();
  if (!q) return [...clients];
  return clients.filter((c) => c.pin.toLowerCase().includes(q));
}}

export type ClientSortField = 'pin' | 'pulse';
export type SortDirection = 'asc' | 'desc';

export function sortClients(
  items: ClientPulse[],
  field: ClientSortField,
  direction: SortDirection,
): ClientPulse[] {{
  const mult = direction === 'asc' ? 1 : -1;

  return [...items].sort((a, b) => {{
    if (field === 'pin') {{
      return mult * a.pin.localeCompare(b.pin);
    }}
    // pulse
    if (a.pulseIndex === null && b.pulseIndex === null) return 0;
    if (a.pulseIndex === null) return 1;
    if (b.pulseIndex === null) return -1;
    return mult * (a.pulseIndex - b.pulseIndex);
  }});
}}
"""

    OUT.write_text(content, encoding="utf-8")
    print(f"Wrote {len(rows)} clients to {OUT}")
    print(f"First PIN: {rows[0]['pin']}")


if __name__ == "__main__":
    main()
